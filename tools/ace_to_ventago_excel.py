#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
ACE 레거시 DB → VentaGO import Excel 변환 스크립트
=============================================================================

오프라인 PC 의 ACE 레거시 PostgreSQL DB 를 읽어,
VentaGO 의 `GET /code-import/template` 형식과 100% 동일한
Excel(.xlsx) 파일을 생성한다. 이 파일을 VentaGO 의
"Importar códigos" 화면에 업로드하면 매장 데이터가 일괄 등록된다.

VentaGO import 템플릿 형식 (api-ventago/src/app/code-import/code-import.controller.ts 기준):
  - Sheet "Colors"        : name, hex
  - Sheet "CodigoMadres"  : sku, name, price, priceOrig, description, categoryName
  - Sheet "CodigoHijitos" : parentSku, sku, name, colorName, size, price,
                            priceOrig, stock, price1, price2, price3, price4, price5

  주의: 템플릿에는 season/origin/supplier 컬럼이 없다.
        ACE 의 temporadas/origenes/empresas 는 매핑할 곳이 없으므로,
        --enrich-description 플래그로 description 에 병합하거나(기본 OFF) 드롭한다.
        store_id 는 지정하지 않는다 — VentaGO import 시 로그인 매장으로 자동 할당.

설계 원칙:
  - 런타임 컬럼 introspection: ACE 스키마 변형(컬럼명 차이)에 견디도록,
    실제 컬럼명을 information_schema 로 탐지 후 후보 목록에서 매핑한다.
  - pool 낭비 방지: psycopg2 connection 1개만 with 블록으로 열고 닫는다.
  - 행 단위 에러 격리: 한 행 실패가 전체 export 를 중단하지 않는다(에러 로그 누적).
  - 디버깅 우선: --debug 로 상세 로그, --inspect 로 스키마만 출력 후 종료.

실행:
  python ace_to_ventago_excel.py --host localhost --port 5432 \\
      --db ace_db --user postgres --output export_ventago.xlsx

환경변수(인자 미지정 시 fallback):
  ACE_DB_HOST, ACE_DB_PORT, ACE_DB_NAME, ACE_DB_USER, ACE_DB_PASSWORD

의존성:
  psycopg2-binary, openpyxl   (requirements.txt 참조)
"""

import argparse
import getpass
import logging
import os
import re
import sys
from decimal import Decimal

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    sys.stderr.write(
        "[FATAL] psycopg2 가 설치되어 있지 않습니다. "
        "`pip install -r requirements.txt` 를 먼저 실행하세요.\n"
    )
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.stderr.write(
        "[FATAL] openpyxl 가 설치되어 있지 않습니다. "
        "`pip install -r requirements.txt` 를 먼저 실행하세요.\n"
    )
    sys.exit(1)


# =============================================================================
# 로깅 설정
# =============================================================================
logger = logging.getLogger("ace2ventago")


def setup_logging(debug: bool) -> None:
    """로깅 초기화 — debug 면 DEBUG, 아니면 INFO 레벨."""
    level = logging.DEBUG if debug else logging.INFO
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    )
    logger.setLevel(level)
    logger.handlers.clear()
    logger.addHandler(handler)


# =============================================================================
# VentaGO 템플릿 헤더 (controller.ts 와 1:1 일치 — 절대 변경 금지)
# =============================================================================
COLORS_HEADER = ["name", "hex"]
MADRES_HEADER = ["sku", "name", "price", "priceOrig", "description", "categoryName"]
HIJITOS_HEADER = [
    "parentSku",
    "sku",
    "name",
    "colorName",
    "size",
    "price",
    "priceOrig",
    "stock",
    "price1",
    "price2",
    "price3",
    "price4",
    "price5",
]


# =============================================================================
# 컬럼 자동 탐지용 후보 목록 (logical field → 실제 컬럼명 후보, 우선순위 순)
# ACE 스키마가 PC 마다 조금씩 다를 수 있어 후보를 넉넉히 둔다.
# 모두 소문자 비교.
# =============================================================================

# --- todocodigos (부모 상품) ---
PARENT_FIELD_CANDIDATES = {
    "sku": ["tcodigo", "codigo", "tcod"],
    "name": ["tdesc", "tdescripcion", "descripcion", "tnombre"],
    "price1": ["tpre1", "tprecio1", "tpvp1", "precio1", "tpre"],
    "price2": ["tpre2", "tprecio2", "tpvp2", "precio2"],
    "price3": ["tpre3", "tprecio3", "tpvp3", "precio3"],
    "price4": ["tpre4", "tprecio4", "tpvp4", "precio4"],
    "price5": ["tpre5", "tprecio5", "tpvp5", "precio5"],
    "deleted": ["borrado", "eliminado", "tborrado", "deleted"],
    # 참조 FK
    "fk_tipo": ["id_tipo", "idtipo", "tipo", "tpcodigo", "ref_id_tipo", "tipos_id"],
    "fk_temporada": [
        "id_temporada",
        "idtemporada",
        "temporada",
        "ref_id_temporada",
        "temporadas_id",
    ],
    "fk_origen": ["id_origen", "idorigen", "origen", "ref_id_origen", "origenes_id"],
    "fk_empresa": [
        "id_empresa",
        "idempresa",
        "empresa",
        "ref_id_empresa",
        "empresas_id",
        "empcodigo",
    ],
}

# --- codigos (자식·변형 상품) ---
VARIANT_FIELD_CANDIDATES = {
    "sku": ["codigo", "ccodigo", "cod"],
    "name": ["descripcion", "cdesc", "cdescripcion", "nombre"],
    "parent_sku": ["codigoproducto", "tcodigo", "codigopadre", "cod_padre", "padre"],
    "price1": ["pre1", "precio1", "pvp1", "cpre1", "pre"],
    "price2": ["pre2", "precio2", "pvp2", "cpre2"],
    "price3": ["pre3", "precio3", "pvp3", "cpre3"],
    "price4": ["pre4", "precio4", "pvp4", "cpre4"],
    "price5": ["pre5", "precio5", "pvp5", "cpre5"],
    "stock": ["stock", "cantidad", "existencia", "existencias", "cstock"],
    "deleted": ["borrado", "eliminado", "cborrado", "deleted"],
    # 색상/사이즈
    "fk_color": ["ref_id_color", "id_color", "idcolor", "color", "colores_id"],
    "size_str": ["str_talle", "talle", "talla", "size", "ctalle"],
}

# --- 참조 테이블: PK 후보 + 표시명(desc) 후보 ---
REFERENCE_TABLES = {
    "color": {
        "pk": ["id", "idcolor", "id_color", "codigo", "ref_id_color"],
        "desc": ["descripcioncolor", "descripcion", "nombre", "color"],
    },
    "tipos": {
        "pk": ["id", "idtipo", "tipo", "tpcodigo", "codigo"],
        "desc": ["tpdesc", "descripcion", "nombre", "tipo_nombre"],
    },
    "temporadas": {
        "pk": ["id", "idtemporada", "temporada", "codigo"],
        "desc": ["temporada_nombre", "nombre", "descripcion", "temporada"],
    },
    "origenes": {
        "pk": ["id", "idorigen", "origen", "codigo"],
        "desc": ["origen_nombre", "nombre", "descripcion", "origen"],
    },
    "empresas": {
        "pk": ["id", "idempresa", "empresa", "empcodigo", "codigo"],
        "desc": ["empdesc", "descripcion", "nombre", "empresa_nombre"],
    },
}

# 색상 약어 → 정식 색상명 (괄호 파싱 fallback 용)
COLOR_ABBREV_MAP = {
    "ROJ": "ROJO",
    "AZU": "AZUL",
    "NEG": "NEGRO",
    "BLA": "BLANCO",
    "VER": "VERDE",
    "AMA": "AMARILLO",
    "NAR": "NARANJA",
    "GRI": "GRIS",
    "MAR": "MARRON",
    "ROS": "ROSA",
    "VIO": "VIOLETA",
    "CEL": "CELESTE",
    "BEI": "BEIGE",
    "FUC": "FUCSIA",
    "BOR": "BORDO",
    "TUR": "TURQUESA",
    "DOR": "DORADO",
    "PLA": "PLATA",
    "CRE": "CREMA",
    "CAM": "CAMEL",
    "LIL": "LILA",
    "MOS": "MOSTAZA",
    "COR": "CORAL",
    "VIN": "VINO",
    "OLI": "OLIVA",
}

# 사이즈 파싱: SKU 끝자리에서 추출할 문자/숫자 사이즈 패턴
ALPHA_SIZES = ["XXXL", "XXL", "XL", "XS", "S", "M", "L"]
# 숫자 사이즈 범위 (의류 일반): 0~16, 짝수 26~60
NUMERIC_SIZE_RE = re.compile(r"(?<![0-9])([0-9]{1,2})$")
PAREN_RE = re.compile(r"\(([^)]+)\)")


# =============================================================================
# 인자 / 환경변수
# =============================================================================
def parse_args() -> argparse.Namespace:
    """CLI 인자 파싱 (환경변수 fallback)."""
    parser = argparse.ArgumentParser(
        description="ACE 레거시 DB → VentaGO import Excel 변환",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--host", default=os.getenv("ACE_DB_HOST", "localhost"), help="DB 호스트"
    )
    parser.add_argument(
        "--port", default=os.getenv("ACE_DB_PORT", "5432"), help="DB 포트"
    )
    parser.add_argument(
        "--db", default=os.getenv("ACE_DB_NAME", "ace_db"), help="DB 이름"
    )
    parser.add_argument(
        "--user", default=os.getenv("ACE_DB_USER", "postgres"), help="DB 유저"
    )
    parser.add_argument(
        "--password",
        default=os.getenv("ACE_DB_PASSWORD"),
        help="DB 비밀번호 (미지정 시 프롬프트 또는 .pgpass)",
    )
    parser.add_argument(
        "--output",
        default="export_ventago.xlsx",
        help="출력 Excel 파일 경로",
    )
    parser.add_argument(
        "--schema",
        default="public",
        help="ACE 테이블이 위치한 스키마",
    )
    parser.add_argument(
        "--include-deleted",
        action="store_true",
        help="borrado=true 인 행도 포함 (기본: 제외)",
    )
    parser.add_argument(
        "--enrich-description",
        action="store_true",
        help="temporada/origen/empresa 를 description 에 병합 (기본: 드롭)",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="DB 스키마(탐지된 컬럼 매핑)만 출력하고 종료 — 디버깅용",
    )
    parser.add_argument(
        "--debug", action="store_true", help="상세(DEBUG) 로그 출력"
    )

    return parser.parse_args()


# =============================================================================
# DB 헬퍼
# =============================================================================
def list_table_columns(cur, schema: str, table: str) -> list:
    """information_schema 로 테이블의 실제 컬럼명 목록(소문자)을 반환. 없으면 빈 리스트."""
    cur.execute(
        """
        SELECT lower(column_name) AS col
        FROM information_schema.columns
        WHERE table_schema = %s AND lower(table_name) = lower(%s)
        ORDER BY ordinal_position
        """,
        (schema, table),
    )
    cols = [r["col"] for r in cur.fetchall()]
    logger.debug("테이블 '%s' 컬럼: %s", table, cols)

    return cols


def detect_column(available_cols: list, candidates: list) -> str:
    """후보 목록에서 실제 존재하는 첫 컬럼명을 반환(소문자). 없으면 None."""
    col_set = set(available_cols)
    for cand in candidates:
        if cand.lower() in col_set:
            return cand.lower()

    return None


def detect_field_map(available_cols: list, field_candidates: dict) -> dict:
    """logical field → 실제 컬럼명 매핑을 한 번에 구성."""
    field_map = {}
    for field, candidates in field_candidates.items():
        col = detect_column(available_cols, candidates)
        field_map[field] = col
        if col is None:
            logger.debug("  미탐지 field='%s' (후보 %s)", field, candidates)
        else:
            logger.debug("  매핑 field='%s' → 컬럼='%s'", field, col)

    return field_map


def load_reference_map(cur, schema: str, table: str, spec: dict) -> dict:
    """
    참조 테이블을 {pk_value(str): desc_value(str)} dict 로 로드.
    PK/desc 컬럼은 후보에서 자동 탐지. 테이블/컬럼 없으면 빈 dict + 경고.
    """
    cols = list_table_columns(cur, schema, table)
    if not cols:
        logger.warning("참조 테이블 '%s' 가 존재하지 않음 — 빈 매핑으로 진행", table)

        return {}

    pk_col = detect_column(cols, spec["pk"])
    desc_col = detect_column(cols, spec["desc"])
    if not pk_col or not desc_col:
        logger.warning(
            "참조 테이블 '%s' PK/desc 컬럼 탐지 실패 (pk=%s, desc=%s) — 빈 매핑",
            table,
            pk_col,
            desc_col,
        )

        return {}

    logger.info(
        "참조 '%s' 로드: pk=%s, desc=%s", table, pk_col, desc_col
    )
    ref_map = {}
    try:
        cur.execute(
            'SELECT "{pk}" AS k, "{desc}" AS v FROM "{schema}"."{table}"'.format(
                pk=pk_col, desc=desc_col, schema=schema, table=table
            )
        )
        for row in cur.fetchall():
            key = row["k"]
            val = row["v"]
            if key is None:
                continue
            ref_map[str(key).strip()] = (val or "").strip() if isinstance(val, str) else val
    except psycopg2.Error as exc:
        logger.warning("참조 '%s' 조회 실패: %s — 빈 매핑", table, exc)

        return {}

    logger.info("참조 '%s' %d 건 로드", table, len(ref_map))

    return ref_map


# =============================================================================
# 값 변환 헬퍼
# =============================================================================
def to_number(value):
    """Decimal/문자/None 을 openpyxl 친화적 숫자(float) 또는 None 으로 변환."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    # 아르헨티나 숫자 형식 처리: "1.500,50" → 1500.50, "1500,50" → 1500.50
    if "," in text and "." in text:
        # 점=천단위, 콤마=소수점
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        # 콤마만 → 소수점
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def clean_str(value) -> str:
    """문자열 정리 — None 은 빈 문자열, 양끝 공백 제거."""
    if value is None:
        return ""

    return str(value).strip()


def parse_color_from_description(desc: str) -> str:
    """
    descripcion 의 괄호 안 텍스트에서 색상 추출.
    예) "REMERA BASICA (ROJ)" → "ROJO"
    약어 매핑에 있으면 정식명, 없으면 괄호 원문(대문자) 반환. 없으면 "".
    """
    if not desc:
        return ""
    matches = PAREN_RE.findall(desc)
    for raw in reversed(matches):  # 보통 마지막 괄호가 색상
        token = raw.strip().upper()
        if not token:
            continue
        # 괄호 안이 사이즈(XL, S, 42 등)면 색상 아님 → 스킵
        if token in ALPHA_SIZES or token.isdigit():
            continue
        if token in COLOR_ABBREV_MAP:
            return COLOR_ABBREV_MAP[token]
        # 약어 3글자가 매핑 키의 접두면 매핑
        if len(token) >= 3 and token[:3] in COLOR_ABBREV_MAP:
            return COLOR_ABBREV_MAP[token[:3]]
        # 괄호 안이 숫자(사이즈)거나 너무 길면 색상 아님
        if token.isdigit() or len(token) > 15:
            continue

        return token

    return ""


def parse_size_from_sku(sku: str) -> str:
    """
    SKU 끝자리에서 사이즈 추출.
    알파 사이즈(XXXL..S) 우선, 그다음 끝 1~2자리 숫자.
    찾지 못하면 "".
    """
    if not sku:
        return ""
    upper = sku.strip().upper()
    # 구분자 뒤 토큰 우선 검사
    tail = re.split(r"[\s\-_./]", upper)[-1] if re.search(r"[\s\-_./]", upper) else upper
    for size in ALPHA_SIZES:
        if tail == size or upper.endswith(size):
            return size
    num = NUMERIC_SIZE_RE.search(tail)
    if num:
        val = int(num.group(1))
        if 0 <= val <= 60:
            return str(val)

    return ""


def build_enriched_description(base_desc, temporada, origen, empresa) -> str:
    """temporada/origen/empresa 를 description 끝에 병합."""
    parts = [clean_str(base_desc)]
    extras = []
    if clean_str(temporada):
        extras.append("Temp: {}".format(clean_str(temporada)))
    if clean_str(origen):
        extras.append("Origen: {}".format(clean_str(origen)))
    if clean_str(empresa):
        extras.append("Prov: {}".format(clean_str(empresa)))
    if extras:
        parts.append("[{}]".format(" | ".join(extras)))

    return " ".join(p for p in parts if p).strip()


# =============================================================================
# 데이터 추출
# =============================================================================
def fetch_parents(cur, schema, args, fmap, refs, stats):
    """
    todocodigos(부모) 조회 → CodigoMadres 행 리스트 반환.
    각 행: dict(헤더 키 → 값). 행 단위 에러는 stats['errors'] 에 누적.
    """
    if not fmap.get("sku"):
        logger.error("todocodigos 에서 sku 컬럼(tcodigo 등)을 못 찾음 — 부모 export 건너뜀")

        return []

    select_cols = sorted(
        {c for c in fmap.values() if c}  # 탐지된 실제 컬럼만 SELECT
    )
    where = ""
    if not args.include_deleted and fmap.get("deleted"):
        where = 'WHERE COALESCE("{}", false) = false'.format(fmap["deleted"])

    sql = 'SELECT {cols} FROM "{schema}"."todocodigos" {where}'.format(
        cols=", ".join('"{}"'.format(c) for c in select_cols),
        schema=schema,
        where=where,
    )
    logger.debug("부모 SQL: %s", sql)
    cur.execute(sql)
    db_rows = cur.fetchall()
    logger.info("todocodigos %d 건 조회", len(db_rows))

    out_rows = []
    for idx, row in enumerate(db_rows):
        try:
            sku = clean_str(row.get(fmap["sku"]))
            if not sku:
                stats["parent_skipped"] += 1
                stats["errors"].append(("CodigoMadres", idx, "EMPTY_SKU", "tcodigo 비어있음"))
                continue

            name = clean_str(row.get(fmap.get("name"))) or sku

            # 카테고리 = tipos.tpdesc (FK lookup)
            category = ""
            if fmap.get("fk_tipo"):
                fk_val = row.get(fmap["fk_tipo"])
                if fk_val is not None:
                    category = clean_str(refs["tipos"].get(str(fk_val).strip(), ""))

            # 부모 가격 = tpre1
            price = to_number(row.get(fmap.get("price1")))

            # description (옵션: temporada/origen/empresa 병합)
            if args.enrich_description:
                temporada = _lookup_ref(refs, "temporadas", row, fmap, "fk_temporada")
                origen = _lookup_ref(refs, "origenes", row, fmap, "fk_origen")
                empresa = _lookup_ref(refs, "empresas", row, fmap, "fk_empresa")
                description = build_enriched_description(name, temporada, origen, empresa)
            else:
                description = ""

            out_rows.append(
                {
                    "sku": sku,
                    "name": name,
                    "price": price,
                    "priceOrig": None,
                    "description": description,
                    "categoryName": category,
                }
            )
            stats["parent_ok"] += 1
        except Exception as exc:  # noqa: BLE001 - 행 격리
            stats["parent_skipped"] += 1
            stats["errors"].append(
                ("CodigoMadres", idx, "ROW_ERROR", str(exc))
            )
            logger.debug("부모 행 %d 처리 실패: %s", idx, exc)

    return out_rows


def _lookup_ref(refs, table, row, fmap, fk_field):
    """FK 컬럼값으로 참조 desc 조회 헬퍼."""
    fk_col = fmap.get(fk_field)
    if not fk_col:
        return ""
    fk_val = row.get(fk_col)
    if fk_val is None:
        return ""

    return clean_str(refs.get(table, {}).get(str(fk_val).strip(), ""))


def fetch_variants(cur, schema, args, fmap, refs, parent_sku_set, stats):
    """
    codigos(자식) 조회 → CodigoHijitos 행 리스트 + 사용된 색상명 set 반환.
    """
    if not fmap.get("sku"):
        logger.error("codigos 에서 sku 컬럼(codigo 등)을 못 찾음 — 자식 export 건너뜀")

        return [], set()

    select_cols = sorted({c for c in fmap.values() if c})
    where = ""
    if not args.include_deleted and fmap.get("deleted"):
        where = 'WHERE COALESCE("{}", false) = false'.format(fmap["deleted"])

    sql = 'SELECT {cols} FROM "{schema}"."codigos" {where}'.format(
        cols=", ".join('"{}"'.format(c) for c in select_cols),
        schema=schema,
        where=where,
    )
    logger.debug("자식 SQL: %s", sql)
    cur.execute(sql)
    db_rows = cur.fetchall()
    logger.info("codigos %d 건 조회", len(db_rows))

    out_rows = []
    used_colors = set()
    for idx, row in enumerate(db_rows):
        try:
            sku = clean_str(row.get(fmap["sku"]))
            if not sku:
                stats["variant_skipped"] += 1
                stats["errors"].append(("CodigoHijitos", idx, "EMPTY_SKU", "codigo 비어있음"))
                continue

            parent_sku = clean_str(row.get(fmap.get("parent_sku")))
            if not parent_sku:
                stats["variant_skipped"] += 1
                stats["errors"].append(
                    ("CodigoHijitos", idx, "EMPTY_PARENT_SKU", "codigoproducto 비어있음")
                )
                continue

            # 부모 존재 여부 경고 (export 는 계속 — import 단계에서 다시 검증됨)
            if parent_sku_set and parent_sku not in parent_sku_set:
                stats["variant_orphan"] += 1
                logger.debug("자식 %s 의 부모 %s 가 todocodigos 에 없음", sku, parent_sku)

            name = clean_str(row.get(fmap.get("name"))) or sku

            # --- 색상 파싱 ---
            color_name = ""
            color_source = ""
            if fmap.get("fk_color"):
                fk_val = row.get(fmap["fk_color"])
                if fk_val is not None and str(fk_val).strip() != "":
                    color_name = clean_str(refs["color"].get(str(fk_val).strip(), ""))
                    if color_name:
                        color_source = "fk"
            if not color_name:
                parsed = parse_color_from_description(clean_str(row.get(fmap.get("name"))))
                if parsed:
                    color_name = parsed
                    color_source = "parsed"
            color_name = color_name.upper() if color_name else ""
            if color_name:
                used_colors.add(color_name)
            else:
                stats["color_fail"] += 1
                stats["color_fail_skus"].append(sku)

            # --- 사이즈 파싱 ---
            size = ""
            size_source = ""
            if fmap.get("size_str"):
                size = clean_str(row.get(fmap["size_str"]))
                if size:
                    size_source = "column"
            if not size:
                parsed_size = parse_size_from_sku(sku)
                if parsed_size:
                    size = parsed_size
                    size_source = "parsed"
            size = size.upper() if size else ""
            if not size:
                stats["size_fail"] += 1
                stats["size_fail_skus"].append(sku)

            logger.debug(
                "자식 %s color=%s(%s) size=%s(%s)",
                sku, color_name, color_source, size, size_source,
            )

            # --- 가격 5종 ---
            p1 = to_number(row.get(fmap.get("price1")))
            p2 = to_number(row.get(fmap.get("price2")))
            p3 = to_number(row.get(fmap.get("price3")))
            p4 = to_number(row.get(fmap.get("price4")))
            p5 = to_number(row.get(fmap.get("price5")))
            stock = to_number(row.get(fmap.get("stock")))

            out_rows.append(
                {
                    "parentSku": parent_sku,
                    "sku": sku,
                    "name": name,
                    "colorName": color_name,
                    "size": size,
                    "price": p1,  # base = price1
                    "priceOrig": None,
                    "stock": stock,
                    "price1": p1,
                    "price2": p2,
                    "price3": p3,
                    "price4": p4,
                    "price5": p5,
                }
            )
            stats["variant_ok"] += 1
        except Exception as exc:  # noqa: BLE001 - 행 격리
            stats["variant_skipped"] += 1
            stats["errors"].append(("CodigoHijitos", idx, "ROW_ERROR", str(exc)))
            logger.debug("자식 행 %d 처리 실패: %s", idx, exc)

    return out_rows, used_colors


def build_colors_rows(refs, used_colors, stats):
    """
    Colors 시트 행 구성:
      - color 테이블의 모든 descripcioncolor (name)
      - + 자식에서 파싱됐지만 테이블에 없던 색상명
    hex 는 ACE 에 없으므로 공란.
    """
    names = set()
    for val in refs.get("color", {}).values():
        name = clean_str(val).upper()
        if name:
            names.add(name)
    table_count = len(names)
    # 파싱으로만 등장한 색상도 추가 (import 시 자동 생성되지만 명시적으로 포함)
    for c in used_colors:
        names.add(clean_str(c).upper())
    parsed_only = len(names) - table_count if len(names) >= table_count else 0
    stats["colors_total"] = len(names)
    stats["colors_from_table"] = table_count
    stats["colors_parsed_only"] = max(0, parsed_only)

    return [{"name": n, "hex": None} for n in sorted(names) if n]


# =============================================================================
# Excel 생성
# =============================================================================
def write_excel(output_path, colors_rows, madres_rows, hijitos_rows):
    """openpyxl 로 3 시트 Excel 생성 (템플릿 헤더 순서 그대로)."""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A2E")  # Ventago 다크 네이비

    def add_sheet(title, header, rows):
        ws = wb.create_sheet(title=title)
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append([row.get(col) for col in header])
        # 컬럼 너비 적당히
        for col_idx, _ in enumerate(header, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16
        ws.freeze_panes = "A2"
        logger.debug("시트 '%s' %d 행 작성", title, len(rows))

    add_sheet("Colors", COLORS_HEADER, colors_rows)
    add_sheet("CodigoMadres", MADRES_HEADER, madres_rows)
    add_sheet("CodigoHijitos", HIJITOS_HEADER, hijitos_rows)

    wb.save(output_path)
    logger.info("Excel 저장 완료: %s", output_path)


# =============================================================================
# inspect 모드
# =============================================================================
def run_inspect(cur, schema):
    """탐지된 스키마/컬럼 매핑을 출력 — 매핑 검증용 디버깅 모드."""
    print("\n===== ACE 스키마 inspect (schema={}) =====".format(schema))
    for table, candidates in (
        ("todocodigos", PARENT_FIELD_CANDIDATES),
        ("codigos", VARIANT_FIELD_CANDIDATES),
    ):
        cols = list_table_columns(cur, schema, table)
        print("\n[{}] 실제 컬럼: {}".format(table, cols or "(테이블 없음)"))
        if cols:
            fmap = detect_field_map(cols, candidates)
            for field, col in fmap.items():
                mark = "OK " if col else "!! "
                print("  {}{:<14} → {}".format(mark, field, col or "(미탐지)"))
    print("\n[참조 테이블]")
    for table, spec in REFERENCE_TABLES.items():
        cols = list_table_columns(cur, schema, table)
        if not cols:
            print("  !! {:<12} (테이블 없음)".format(table))
            continue
        pk = detect_column(cols, spec["pk"])
        desc = detect_column(cols, spec["desc"])
        print("  OK {:<12} pk={}, desc={}".format(table, pk, desc))
    print("\n=========================================\n")


# =============================================================================
# 리포트
# =============================================================================
def print_report(stats, output_path):
    """결과 요약 리포트 출력."""
    print("\n" + "=" * 60)
    print("  ACE → VentaGO Excel 변환 완료")
    print("=" * 60)
    print("  출력 파일       : {}".format(output_path))
    print("-" * 60)
    print("  색상(Colors)    : {} 건 (테이블 {} + 파싱전용 {})".format(
        stats.get("colors_total", 0),
        stats.get("colors_from_table", 0),
        stats.get("colors_parsed_only", 0),
    ))
    print("  부모(CodigoMadres) : {} 건 export, {} 건 스킵".format(
        stats["parent_ok"], stats["parent_skipped"]
    ))
    print("  자식(CodigoHijitos): {} 건 export, {} 건 스킵".format(
        stats["variant_ok"], stats["variant_skipped"]
    ))
    print("-" * 60)
    print("  ⚠ 색상 파싱 실패 : {} 건 (수동 확인 필요)".format(stats["color_fail"]))
    print("  ⚠ 사이즈 파싱 실패: {} 건 (수동 확인 필요)".format(stats["size_fail"]))
    print("  ⚠ 부모 없는 자식  : {} 건 (import 시 해당 행 거부됨)".format(
        stats["variant_orphan"]
    ))
    print("  ⚠ 행 처리 에러    : {} 건".format(len(stats["errors"])))

    # 실패 SKU 일부 샘플 출력 (최대 15개)
    if stats["color_fail_skus"]:
        print("\n  [색상 파싱 실패 SKU 샘플]")
        print("   " + ", ".join(stats["color_fail_skus"][:15]))
    if stats["size_fail_skus"]:
        print("\n  [사이즈 파싱 실패 SKU 샘플]")
        print("   " + ", ".join(stats["size_fail_skus"][:15]))
    if stats["errors"]:
        print("\n  [행 에러 샘플 (최대 10)]")
        for sheet, idx, code, msg in stats["errors"][:10]:
            print("   - {} row#{} [{}] {}".format(sheet, idx, code, msg))
    print("=" * 60 + "\n")


# =============================================================================
# main
# =============================================================================
def main():
    args = parse_args()
    setup_logging(args.debug)

    password = args.password
    if password is None and sys.stdin.isatty():
        # 비밀번호 미지정 + 인터랙티브면 안전하게 프롬프트
        try:
            password = getpass.getpass("ACE DB 비밀번호 (없으면 Enter): ") or None
        except (EOFError, KeyboardInterrupt):
            password = None

    conn_kwargs = {
        "host": args.host,
        "port": args.port,
        "dbname": args.db,
        "user": args.user,
    }
    if password:
        conn_kwargs["password"] = password

    logger.info(
        "ACE DB 연결 시도: host=%s port=%s db=%s user=%s",
        args.host, args.port, args.db, args.user,
    )

    stats = {
        "parent_ok": 0,
        "parent_skipped": 0,
        "variant_ok": 0,
        "variant_skipped": 0,
        "variant_orphan": 0,
        "color_fail": 0,
        "size_fail": 0,
        "color_fail_skus": [],
        "size_fail_skus": [],
        "errors": [],
    }

    # pool 낭비 방지: connection 1개만 with 블록으로 — 종료 시 자동 close
    try:
        with psycopg2.connect(**conn_kwargs) as conn:
            conn.set_session(readonly=True, autocommit=True)  # 읽기 전용 보장
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                # --- inspect 모드: 매핑만 출력 후 종료 ---
                if args.inspect:
                    run_inspect(cur, args.schema)

                    return 0

                # --- 참조 테이블 로드 ---
                logger.info("참조 테이블 로드 중...")
                refs = {}
                for table, spec in REFERENCE_TABLES.items():
                    refs[table] = load_reference_map(cur, args.schema, table, spec)

                # --- 부모(todocodigos) ---
                parent_cols = list_table_columns(cur, args.schema, "todocodigos")
                if not parent_cols:
                    logger.error("todocodigos 테이블이 없음 — 중단")

                    return 2
                parent_fmap = detect_field_map(parent_cols, PARENT_FIELD_CANDIDATES)
                madres_rows = fetch_parents(
                    cur, args.schema, args, parent_fmap, refs, stats
                )
                parent_sku_set = {r["sku"] for r in madres_rows}

                # --- 자식(codigos) ---
                variant_cols = list_table_columns(cur, args.schema, "codigos")
                if not variant_cols:
                    logger.error("codigos 테이블이 없음 — 중단")

                    return 2
                variant_fmap = detect_field_map(variant_cols, VARIANT_FIELD_CANDIDATES)
                hijitos_rows, used_colors = fetch_variants(
                    cur, args.schema, args, variant_fmap, refs, parent_sku_set, stats
                )

                # --- 색상 시트 ---
                colors_rows = build_colors_rows(refs, used_colors, stats)

    except psycopg2.OperationalError as exc:
        logger.error("DB 연결 실패: %s", exc)
        logger.error(
            "호스트/포트/DB명/유저/비밀번호를 확인하세요. "
            "(환경변수 ACE_DB_* 또는 --host/--port/--db/--user/--password)"
        )

        return 3
    except psycopg2.Error as exc:
        logger.error("DB 오류: %s", exc)

        return 4

    # --- Excel 생성 ---
    try:
        write_excel(args.output, colors_rows, madres_rows, hijitos_rows)
    except Exception as exc:  # noqa: BLE001
        logger.error("Excel 생성 실패: %s", exc)

        return 5

    print_report(stats, args.output)

    return 0


if __name__ == "__main__":
    sys.exit(main())
